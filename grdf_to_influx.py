from openpyxl import load_workbook
import glob
from datetime import datetime
from influxdb import InfluxDBClient

export_files = [f for f in glob.glob("Consommations gaz_*.xlsx")]
# Let's take the first found
filename = export_files[0]
#Extract the date range
format = '%d%m%Y'
first_date = datetime.strptime(filename.split('_')[2], format)
last_date = datetime.strptime(filename.split('_')[3][:-5], format)
total_days = (last_date - first_date).days

#Load the excel file
wb = load_workbook(filename=export_files[0], read_only=True)
ws = wb['Historique par jour']

format = "%d/%m/%Y"
json = []
for row in ws.iter_rows(min_row=8, max_row=8+total_days, max_col=6, values_only=True):
    date = row[1]
    m3 = row[4]
    kwh = row[5]
    print(f'Adding the consumption for the {date}: {m3} m³ and {kwh} kWh.')
    json.append({
        "measurement": "conso_gaz",
        "tags": {
            "fetch_date": datetime.today().date().strftime(format)
        },
        "time": datetime.strptime(date + " 12:00", format + " %H:%M").strftime('%Y-%m-%dT%H:%M:%SZ'),
        "fields": {
            "kwh": kwh,
            "mcube": m3
        }
    })

client = InfluxDBClient(host='192.168.1.99', port=8086)
client.create_database('gazpar')
client.switch_database('gazpar')
client.write_points(json)
